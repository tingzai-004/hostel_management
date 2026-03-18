from django.shortcuts import render,HttpResponse,redirect
from . import models
from hostel import models
from hostel.models import area ,dorm
from django import forms
from openpyxl import load_workbook
from hostel.bootstrap import BootStrapModelForm
import logging
from django.core.paginator import Paginator,PageNotAnInteger,EmptyPage
from django.db.models import Count,Sum
from django.contrib import messages
from hostel.person import parse_date
from decimal import Decimal,DecimalException
import decimal
logger=logging.getLogger('management')

class feetypemodel(BootStrapModelForm):
    bootstrap_exclude=["is_sign"]
    class Meta:
        model=models.feetype
        fields=['name','bit','cdate','status','fee','area',"effective_date"]
def fee_type(request):#计费类型
    search_data = request.GET.get('q', '')
    query = models.feetype.objects.all()
    if search_data:
        query = query.filter(area__name__contains=search_data )  # 搜索过滤

    # 2. 分页配置（每页5条）
    page_number = request.GET.get('page', 1)
    paginator = Paginator(query, 5)
    page_obj = paginator.get_page(page_number)
    current = page_obj.number
    total_pages = paginator.num_pages

    # 核心算法：始终显示当前页前后各1页，共3个（边界处理）
    if total_pages <= 3:
        display_pages = list(range(1, total_pages + 1))
    else:
        if current <= 2:
            display_pages = [1, 2, 3]
        elif current >= total_pages - 1:
            display_pages = [total_pages - 2, total_pages - 1, total_pages]
        else:
            display_pages = [current - 1, current, current + 1]
    try:
        c_page = paginator.page(page_number)
    except PageNotAnInteger:
        c_page = paginator.page(1)
    except EmptyPage:
        c_page = paginator.page(paginator.num_pages)

    # 4. 传数据到模板
    return render(request, 'fee_type.html', {
        'c_page': c_page,
        'paginator': paginator,
        'search_data': search_data,
        "titel":"费用类型",
        "display_pages":display_pages,
    })
##1  

def add_feetype(request):
    if request.method == "GET":
        form = feetypemodel()
        return render(request, "add_area.html", {"form": form, "titel": "添加计费类型"})
    
    form = feetypemodel(data=request.POST)
    if form.is_valid():
        # feetype_obj = form.save()  # 保存新的计费类型
        # # 新计费类型需要触发所属区域所有房间的当月费用计算
        # rooms = Room.objects.filter(dorm__area=feetype_obj.area)  # 该区域下所有房间
        # current_year = feetype_obj.cdate.year  # 用创建时间的年月
        # current_month = feetype_obj.cdate.month
        
        # for room in rooms:
        #     recalculate_fee_and_sharing(room, current_year, current_month)
        form.save()
        
        return redirect("/hostel/fee_type/")
    return render(request, "add_area.html", {"form": form, "error": form.errors, "titel": "添加计费类型"})

def del_feetype(request,id):
    row_object=models.feetype.objects.filter(id=id).first()
    if row_object==None:
        return HttpResponse("数据不存在")
    # data=request.session.get('info')
    # if data==None:
    #     return HttpResponse("请先登录")
    # name=data.get('name')
    # logger.info('%s删除了部门数据'%(name))
    row_object.delete()
    return redirect("/hostel/fee_type/")

# def edit_feetype(request,id):
#     row_object=models.feetype.objects.filter(id=id).first()
#     if row_object==None:
#         return HttpResponse("要修改的数据不存在")
#     if request.method=='GET':
#         titel="修改-{}".format(row_object.name)
#         form=feetypemodel(instance=row_object)
#         return render(request,'add_area.html',{"form":form,"titel":titel})
#     form=feetypemodel(data=request.POST,instance=row_object)
#     if form.is_valid():
#         # data=request.session.get('info')
#         # if data==None:
#         #     return HttpResponse("请先登录")
#         # name=data.get('name')
#         # logger.info('%s更新了部门数据'%(name))
#         form.save()
#         return redirect("/hostel/fee_type/")
#     return render(request,'add_area.html',{"form":form,"error":form.errors,"titel":titel})

def edit_feetype(request, id):
    feetype_obj = get_object_or_404(models.feetype, id=id)
    if request.method == 'GET':
        titel = f"修改-{feetype_obj.name}"
        form = feetypemodel(instance=feetype_obj)
        return render(request, 'add_area.html', {"form": form, "titel": titel})
    
    form = feetypemodel(data=request.POST, instance=feetype_obj)
    if form.is_valid():
        # updated_feetype = form.save()  # 保存修改后的计费类型（单价/状态已变更）
        
        # # 关键：触发该区域下所有房间的费用重新计算
        # area = updated_feetype.area
        # rooms = Room.objects.filter(dorm__area=area)  # 该区域所有房间
        
        # for room in rooms:
        #     # 取该房间最近3个月的费用记录（可调整数量）
        #     recent_fee_records = fee_record_one.objects.filter(room=room).order_by('-date')[:3]
        #     for fee_record in recent_fee_records:
        #         recalculate_fee_and_sharing(
        #             room=room,
        #             year=fee_record.date.year,
        #             month=fee_record.date.month
        #         )
        form.save()
        return redirect("/hostel/fee_type/")
    return render(request, 'add_area.html', {"form": form, "error": form.errors, "titel": titel})



def fee_recode(request):#费用细则
    
    search_data = request.GET.get('q', '')
    query = models.fee_record_one.objects.all()
    if search_data:
        query = query.filter(room__room_name__contains=search_data)

    
    unique_fee_names = models.feetype.objects.filter(status="启用").values_list('name', flat=True).distinct()
    active_fees = [{'name': name} for name in unique_fee_names]

    # 分页配置（不变）
    page_number = request.GET.get('page', 1)
    paginator = Paginator(query, 5)
    page_obj = paginator.get_page(page_number)
    current = page_obj.number
    total_pages = paginator.num_pages

    # 核心算法：始终显示当前页前后各1页，共3个（边界处理）
    if total_pages <= 3:
        display_pages = list(range(1, total_pages + 1))
    else:
        if current <= 2:
            display_pages = [1, 2, 3]
        elif current >= total_pages - 1:
            display_pages = [total_pages - 2, total_pages - 1, total_pages]
        else:
            display_pages = [current - 1, current, current + 1]
    try:
        c_page = paginator.page(page_number)
    except PageNotAnInteger:
        c_page = paginator.page(1)
    except EmptyPage:
        c_page = paginator.page(paginator.num_pages)
    return render(request, 'fee_recode.html', {
        'c_page': c_page,
        'paginator': paginator,
        'search_data': search_data,
        "titel":"宿舍缴费记录",
        "active_fees":active_fees,
        'display_pages': display_pages,  # 去重后的表头
    })



class fee_recordmodel(BootStrapModelForm):
    class Meta:
        model=models.fee_record_one
        fields=['status']
        widgets={
            'date':forms.TextInput(attrs={'placeholder':'如2025-11'})
        }

##增加删除
##1
def add_fee_record(request):
    return render(request,'add_area.html')
    


   

def del_fee_record(request,id):
    row_object=models.fee_record_one.objects.filter(id=id).first()
    if row_object==None:
        return HttpResponse("数据不存在")
    data=request.session.get('info')
    if data==None:
        return HttpResponse("请先登录")
    name=data.get('name')
    logger.info('%s删除了宿舍费用记录数据'%(name))
    row_object.delete()
    return redirect("/hostel/fee_record?page={}".format(request.GET.get("page",1)))

#更新和批量添加
def edit_fee_record(request,id):
    row_object=models.fee_record_one.objects.filter(id=id).first()
    if row_object==None:
        return HttpResponse("要修改的数据不存在")
    if request.method=='GET':
        titel="修改-{}".format(row_object.room)
        form=fee_recordmodel(instance=row_object)
        return render(request,'add_area.html',{"form":form,"titel":titel})
    form=fee_recordmodel(data=request.POST,instance=row_object)
    if form.is_valid():
        # data=request.session.get('info')
        # if data==None:
        #     return HttpResponse("请先登录")
        # name=data.get('name')
        # logger.info('%s更新了部门数据'%(name))
        form.save()
        return redirect("/hostel/fee_record/")
    return render(request,'add_area.html',{"form":form,"error":form.errors,"titel":titel})




#批量上传
def  add_all_fee_record(request):
       
    if request.method == 'GET':
        return render(request, 'add_all_area.html')

    excel_file = request.FILES.get('file')
    if not excel_file:
        return HttpResponse("请选择文件")

    wb = load_workbook(excel_file)
    sheet = wb.active
    errors = []  # 收集错误信息

    # 假设Excel列顺序：
    # row[0] = 用户名（关联Person的name）
    # row[1] = （预留或其他字段，根据实际情况调整）
    # row[2] = 房间名称（关联Room的room_name）
    # row[3] = 入住日期（check_in_date）
    # row[4] = 退房日期（check_out_date）
    # row[5] = 状态（status）

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # 1. 解析用户（外键：通过姓名关联Person）
            date = row[0]
            
            if not date:
                errors.append(f"第{idx}行错误：状态为空")
                continue  # 跳行
            date= str(date).strip() 

            # 2. 解析房间（外键：通过房间名称关联Room）
            amount = row[1]
            if not amount:
                errors.append(f"第{idx}行错误：状态为空")
                continue  # 跳行
            amount = str(amount).strip() 

            status = row[2]
            if not status:
                errors.append(f"第{idx}行错误：日期为空")
                continue  # 跳行
            status = str(status).strip() 


            # 3. 解析入住日期（调用辅助函数，处理空值和格式）
            try:
                cdate = parse_date(row[3])
            except ValueError as e:
                errors.append(f"第{idx}行错误：入住日期解析失败 - {str(e)}")
                continue  # 跳行
            check_people = row[4]
            if not check_people:
                errors.append(f"第{idx}行错误：状态为空")
                continue  # 跳行
            check_people = str(check_people).strip()  # 转为字符串并去空格

            room=row[5]
            if not room:
                errors.append(f"第{idx}行错误：状态为空")
                continue  # 跳行
            room=models.Room.objects.filter(id=room).first()
            if not room:
                errors.append(f"第{idx}行错误：房间 '{room}' 不存在")
                continue  # 跳行

           
            models.fee_record.objects.create(
                date=date,
                amount=amount,
                cdate=cdate,
                room=room,
                status=status,
                feetype_id=fee_type,
            )

        except Exception as e:
            errors.append(f"第{idx}行错误：{str(e)}")  # 捕获其他意外错误

    # 处理结果
    if errors:
        return HttpResponse("导入失败：\n" + "\n".join(errors))
    
    logger.info(f"{request.session.get('info', {}).get('name')}批量添加入住记录")
    return redirect('/hostel/fee_record/')
##批量删除
def delete_all_fee_record(request):
    if request.method=="POST":
        
        ids=request.POST.getlist("ids")
        if not ids:
            messages.error(request,"请选择要删除人员")
            return redirect("/hostel/fee_record/")
        try:
            models.fee_record_one.objects.filter(id__in=ids).delete()##delete少个括号不能删除
            logger.info(f"管理员{request.session.get('info', {}).get('name')}批量删除了记录")
        
            messages.success(request,f"成功删除{len(ids)}条数据")
        except Exception as e:
            messages.error(request,"删除失败")
        return redirect("/hostel/fee_record/")
    if request.method=="GET":
        return redirect("/hostel/fee_record/")


   


       

from django.shortcuts import render
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import F
from . import models


class feesharingmodel(BootStrapModelForm):
    class Meta:
        model=models.sharing
        fields=["user","fee_record"]

def sharing(request):
    search_data = request.GET.get('q', '')
    query = models.sharing.objects.all()
    if search_data:
        query = query.filter(user__occupancyrecord__name__contains=search_data)
    unique_fee_names = models.feetype.objects.filter(status="启用").values_list('name', flat=True).distinct()
    active_fees = [{'name': name} for name in unique_fee_names]

    # 分页配置（不变）
    page_number = request.GET.get('page', 1)
    paginator = Paginator(query, 5)
    page_obj = paginator.get_page(page_number)
    current = page_obj.number
    total_pages = paginator.num_pages

    # 核心算法：始终显示当前页前后各1页，共3个（边界处理）
    if total_pages <= 3:
        display_pages = list(range(1, total_pages + 1))
    else:
        if current <= 2:
            display_pages = [1, 2, 3]
        elif current >= total_pages - 1:
            display_pages = [total_pages - 2, total_pages - 1, total_pages]
        else:
            display_pages = [current - 1, current, current + 1]
    try:
        c_page = paginator.page(page_number)
    except PageNotAnInteger:
        c_page = paginator.page(1)
    except EmptyPage:
        c_page = paginator.page(paginator.num_pages)
    return render(request, 'feesharing.html', {
        'c_page': c_page,
        'paginator': paginator,
        'search_data': search_data,
        "titel":"宿舍缴费记录",
        "active_fees":active_fees,
        'display_pages': display_pages,  # 去重后的表头
    })


def del_feesharing(request,id):
    row_object=models.sharing.objects.filter(id=id).first()
    if row_object==None:
        return HttpResponse("数据不存在")
    # data=request.session.get('info')
    # if data==None:
    #     return HttpResponse("请先登录")
    # name=data.get('name')
    # logger.info('%s删除了部门数据'%(name))
    row_object.delete()
    return redirect("/hostel/feesharing/?page={}".format(request.GET.get("page",1)))

def edit_feesharing(request,id):
    row_object=models.sharing.objects.filter(id=id).first()
    if row_object==None:
        return HttpResponse("要修改的数据不存在")
    if request.method=='GET':
        titel="修改-{}".format(row_object.user)
        form=feesharingmodel(instance=row_object)
        return render(request,'add_area.html',{"form":form,"titel":titel})
    form=feesharingmodel(data=request.POST,instance=row_object)
    if form.is_valid():
        # data=request.session.get('info')
        # if data==None:
        #     return HttpResponse("请先登录")
        # name=data.get('name')
        # logger.info('%s更新了部门数据'%(name))
        form.save()
        return redirect("/hostel/feesharing/")
    return render(request,'add_area.html',{"form":form,"error":form.errors,"titel":titel})

##批量上传
def add_all_feesharing(request):
    if request.method == 'GET':
        return render(request, 'add_all_area.html')

    excel_file = request.FILES.get('file')
    if not excel_file:
        return HttpResponse("请选择文件")

    wb = load_workbook(excel_file)
    sheet = wb.active
    errors = []

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # 1. 解析费用细则（外键：fee_record）
            fee_record_name = row[0]
            if not fee_record_name:
                errors.append(f"第{idx}行错误：费用细则为空")
                continue
            fee_record = models.fee_record.objects.filter(name=fee_record_name).first()
            if not fee_record:
                errors.append(f"第{idx}行错误：费用细则 '{fee_record_name}' 不存在")
                continue

            # 2. 解析计费类型（外键：feetype）
            feetype_name = row[1]
            if not feetype_name:
                errors.append(f"第{idx}行错误：计费类型为空")
                continue
            feetype = models.feetype.objects.filter(name=feetype_name).first()
            if not feetype:
                errors.append(f"第{idx}行错误：计费类型 '{feetype_name}' 不存在")
                continue

            # 3. 解析住宿记录（外键：occupancyrecord）
            occupancyrecord_name = row[2]
            if not occupancyrecord_name:
                errors.append(f"第{idx}行错误：住宿记录为空")
                continue
            occupancyrecord = models.occupancyrecord.objects.filter(name=occupancyrecord_name).first()
            if not occupancyrecord:
                errors.append(f"第{idx}行错误：住宿记录 '{occupancyrecord_name}' 不存在")
                continue

            # 4. 解析个人分摊金额
            feesharing_amount = row[3]
            if not feesharing_amount:
                errors.append(f"第{idx}行错误：个人分摊金额为空")
                continue
            feesharing_amount = str(feesharing_amount).strip()

            # 5. 解析是否缴纳（状态）
            status = row[4]
            if status not in ['0', '1']:
                errors.append(f"第{idx}行错误：是否缴纳状态必须为 '0' 或 '1'，当前值为 '{status}'")
                continue

            # 6. 解析开始计费日期
            start_date = parse_date(row[5])
            if not start_date:
                errors.append(f"第{idx}行错误：开始计费日期格式错误（需为 YYYY-MM-DD）")
                continue

            # 7. 解析结束计费日期
            end_date = parse_date(row[6])
            if not end_date:
                errors.append(f"第{idx}行错误：结束计费日期格式错误（需为 YYYY-MM-DD）")
                continue

            # 8. 解析缴纳时间（可为空）
            pay_date = parse_date(row[7]) if row[7] else None

            # 9. 创建费用分摊记录
            models.feesharing.objects.create(
                fee_record=fee_record,
                fee_type=feetype,
                occupancyrecord=occupancyrecord,
                feesharing=feesharing_amount,
                status=status,
                start_date=start_date,
                end_date=end_date,
                date=pay_date
            )

        except Exception as e:
            errors.append(f"第{idx}行错误：{str(e)}")

    if errors:
        return HttpResponse("导入失败：\n" + "\n".join(errors))
    
    logger.info(f"{request.session.get('info', {}).get('name')}批量添加费用分摊记录")
    messages.success(request, "批量导入成功！")
    return redirect('/hostel/feesharing/')


def del_all_feesharing(request):#批量删除
    if request.method=="POST":
        ids=request.POST.getlist("ids")
        if not ids:
            messages.error(request,"请选择要删除的对象")
        try:
            models.feesharing.objects.filter(id_in=ids).delete()
            messages.success(request,f"成功删除{len(ids)}条数据")
            logger.info(f"管理员{request.session.get('info', {} ).get('name')}删除了{len(ids)}条分摊记录")
        except Exception as e:
            messages.error(request,'删除失败')
        return redirect("/hostel/feesharing/")
    return redirect("/hostel/feesharing/")



# views.py（只保留以下核心代码，删除其他冗余）
from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from decimal import Decimal
# from .models import fee_record_one, feetype, resource_useage, Room


# 1. 费用记录列表页（含动态表头、搜索、分页）

# 2. 新增费用记录（自动计算动态费用）


# 3. 删除费用记录（单个删除）
# def del_fee_record(request, id):
#     # 获取要删除的记录，不存在则返回404
#     record = get_object_or_404(fee_record_one, id=id)
#     record.delete()
#     # 删除后跳回列表页（保留当前页码）
#     page = request.GET.get('page', 1)
#     return redirect(f"/hostel/fee_record/?page={page}")





