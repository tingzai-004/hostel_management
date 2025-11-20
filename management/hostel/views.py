from django.shortcuts import render,HttpResponse,redirect
from . import models
from hostel import models
from hostel.models import area ,dorm
from django import forms
from openpyxl import load_workbook
from hostel.bootstrap import BootStrapModelForm
import logging
from django.core.paginator import Paginator, EmptyPage,PageNotAnInteger
import io
from django.contrib import messages


logger=logging.getLogger('management')
#区域
class areamodel(forms.ModelForm):
    class Meta:
        model=area
        fields=['name']

def area(request):
    # 1. 查数据（带搜索）
    search_data = request.GET.get('q', '')
    query = models.area.objects.all()
    if search_data:
        query = query.filter(name__contains=search_data)  # 搜索过滤

    # 2. 分页配置（每页5条）
    paginator = Paginator(query, 5)
    page_num = request.GET.get('page', 1)  # 拿页码，默认1

    # 3. 关键：容错处理（不管页码错成啥样，都返回有效页）
    try:
        c_page = paginator.page(page_num)  # 尝试拿指定页
    except PageNotAnInteger:  # 页码不是数字（比如?page=abc）
        c_page = paginator.page(1)  # 给第1页
    except EmptyPage:  # 页码超出范围（比如总1页，?page=2）
        c_page = paginator.page(paginator.num_pages)  # 给最后1页

    # 4. 传数据到模板
    return render(request, 'area.html', {
        'c_page': c_page,
        'paginator': paginator,
        'search_data': search_data
    })
   

def add_area(request):
    if request.method=="GET":
        title="添加区域"
        form=areamodel()
        return render(request,'add_area.html',{'form':form,'titel':title})
    form=areamodel(data=request.POST)
    if form.is_valid():
        data=request.session.get('info')
        if data==None:
            return HttpResponse("请先登录")
        name=data.get('name')
        logger.info('%s添加了区域%s'%(name,form.cleaned_data.get('name')))

        form.save()
        return redirect('/hostel/area/')
    else:
        return render(request,'add_area.html',{'form':form,'error':form.errors})

    
def delete_area(request,id):
    data=request.session.get('info')
    if data==None:
            return HttpResponse("请先登录")
    name=data.get('name')
    page_num=request.GET.get("page",1)
    
    page_num=int(page_num)
    
    logger.info(f'{name}删除了区域一条记录')
    models.area.objects.filter(id=id).delete()
    return redirect(f'/hostel/area/?page={page_num}')



def edit_area(request,id):
    row_obj=models.area.objects.filter(id=id).first()
    if request.method=="GET":
        title='重置-{}'.format(row_obj.name)
        form=areamodel(instance=row_obj)
        return render(request,'add_area.html',{'form':form,'titel':title})
    form=areamodel(data=request.POST,instance=row_obj)
    
    if form.is_valid():
        form.save()
        return redirect('/hostel/area/')
    else:
        return render(request,'add_area.html',{'form':form,'error':form.errors})

def add_all_area(request):#批量上传
    if request.method=='GET':
        return render(request,'add_all_area.html')
    row_object=request.FILES.get('file')
    wb= load_workbook(row_object)
    sheet=wb.worksheets[0]
    data=request.session.get('info')
    if data==None:
            return HttpResponse("请先登录")
    name=data.get('name')
    logger.info('%s批量添加了区域数据%s'%(name))
    for file in sheet.iter_rows(min_row=2):
        name=file[0].value
        models.area.objects.create(name=name)
    
    return redirect('/hostel/area/')

def delete_all_area(request):
    if request.method == 'POST':  # ① 判断请求是否为POST（只有提交表单时才会进入这个分支）
        ids = request.POST.getlist('ids')  # ② 获取前端勾选的人员ID列表（name="ids"的复选框值）
        if not ids:  # ③ 如果没有选中任何ID，给出错误提示
            messages.error(request, "请选择要删除的人员")
            return redirect('/hostel/area/')  # 重定向回人员列表页
        
        try:
            # ④ 批量删除选中的人员（根据ID列表过滤并删除）
            models.area.objects.filter(id__in=ids).delete()
            messages.success(request, f"成功删除 {len(ids)} 条人员记录")  # ⑤ 成功提示
        except Exception as e:  # ⑥ 捕获删除过程中可能出现的异常（如外键关联错误）
            messages.error(request, f"删除失败：{str(e)}")
        
        return redirect('/hostel/area/')  # ⑦ 无论成功或失败，都重定向回人员列表页
    
    # 如果是GET请求（直接访问这个URL），也重定向回列表页
    return redirect('person_list')

    
    
    


# Create your views here.
#楼栋
class dormmodel(BootStrapModelForm):
    class Meta:
        model=dorm
        fields=['name','area']
        bootstrap_exclude=['']
        
def dorm(request):
    search_data = request.GET.get('q', '')
    query = models.dorm.objects.all()
    if search_data:
        query = query.filter(name__contains=search_data)  # 搜索过滤

    # 2. 分页配置（每页5条）
    paginator = Paginator(query, 5)
    page_num = request.GET.get('page', 1)  # 拿页码，默认1

    # 3. 关键：容错处理（不管页码错成啥样，都返回有效页）
    try:
        c_page = paginator.page(page_num)  # 尝试拿指定页
    except PageNotAnInteger:  # 页码不是数字（比如?page=abc）
        c_page = paginator.page(1)  # 给第1页
    except EmptyPage:  # 页码超出范围（比如总1页，?page=2）
        c_page = paginator.page(paginator.num_pages)  # 给最后1页

    # 4. 传数据到模板
    return render(request, 'dorm.html', {
        'c_page': c_page,
        'paginator': paginator,
        'search_data': search_data
    })

def add_dorm(request):
    if request.method=="GET":
        title="添加楼栋"
        form=dormmodel()
        return render(request,"add_area.html",{"form":form,"titel":title})
    form=dormmodel(data=request.POST)
    if form.is_valid():
        data=request.session.get('info')
        if data==None:
            return HttpResponse("请先登录")
        name=data.get('name')
        logger.info('%s添加了楼栋%s'%(name,form.cleaned_data.get('name')))

        form.save()
        return redirect("/hostel/dorm/")
    return render(request,"add_area.html",{"form":form,"error":form.errors,'titel':"添加楼栋"})

def delete_dorm(request,id):
    page_num=request.GET.get("page","")
    page_num=int(page_num)
    row_object=models.dorm.objects.filter(id=id).first()
    if row_object==None:
        return HttpResponse("要删除的数据不存在")
    data=request.session.get('info')
    if data==None:
            return HttpResponse("请先登录")
    name=data.get('name')
    logger.info('%s删除了一条楼栋表记录'%(name))

    row_object.delete()
    return redirect(f"/hostel/dorm/?page={page_num}")

def edit_dorm(request,id):#编辑
    row_object=models.dorm.objects.filter(id=id).first()
    if row_object==None:
        return HttpResponse("要修改的数据不存在")
    if request.method=="GET":
        title="修改-{}".format(row_object.name)
        form=dormmodel(instance=row_object)
        return render(request,"add_area.html",{"form":form,"titel":title})
    form=dormmodel(data=request.POST,instance=row_object)
    if form.is_valid():
        data=request.session.get('info')
        if data==None:
            return HttpResponse("请先登录")
        name=data.get('name')
        logger.info('%s更新了楼栋信息%s'%(name,form.cleaned_data.get('name')))
        form.save()
        return redirect("/hostel/dorm/")
    return render(request,"add_area.html",{"form":form,"error":form.errors})

def add_all_dorm(request):#批量上传
    if request.method == 'GET':
        return render(request, 'add_all_area.html')

    excel_file = request.FILES.get('file')
    if not excel_file:
        return HttpResponse("请选择文件")

    wb = load_workbook(excel_file)
    sheet = wb.active
    errors = []  # 收集错误信息

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # 解析数据（按Excel列顺序：name, phone, depart_id, gender_id, permission_id, room_id, user_status_id）
            data = {
                'name': str(row[0]).strip() if row[0] else f"未知{idx}",   
                'area': models.area.objects.get(name=row[1]),
            
            }
            models.dorm.objects.create(**data)  # 解包创建
        except Exception as e:
            errors.append(f"第{idx}行错误：{str(e)}")  # 记录错误行

    # 处理结果
    if errors:
        return HttpResponse("导入失败：\n" + "\n".join(errors))
    
    logger.info(f"{request.session.get('info', {}).get('name')}批量添加人员")
    return redirect('/hostel/dorm/')

#批量删除
def delete_all_dorm(request):
    if request.method == 'POST':  # ① 判断请求是否为POST（只有提交表单时才会进入这个分支）
        ids = request.POST.getlist('ids')  # ② 获取前端勾选的人员ID列表（name="ids"的复选框值）
        if not ids:  # ③ 如果没有选中任何ID，给出错误提示
            messages.error(request, "请选择要删除的人员")
            return redirect('/hostel/dorm/')  # 重定向回人员列表页
        
        try:
            # ④ 批量删除选中的人员（根据ID列表过滤并删除）
            models.dorm.objects.filter(id__in=ids).delete()
            messages.success(request, f"成功删除 {len(ids)} 条人员记录")  # ⑤ 成功提示
        except Exception as e:  # ⑥ 捕获删除过程中可能出现的异常（如外键关联错误）
            messages.error(request, f"删除失败：{str(e)}")
        
        return redirect('/hostel/dorm/')  # ⑦ 无论成功或失败，都重定向回人员列表页
    
    # 如果是GET请求（直接访问这个URL），也重定向回列表页
    return redirect('/hostel/dorm/')


    
##房间
class roommodel(BootStrapModelForm):
    class Meta:
        model=models.Room
        fields=['room_name','people','room_status','dorm','gender']

def room(request):
   
    search_data = request.GET.get('q', '')
    query = models.Room.objects.all()
    if search_data:
        query = query.filter(room_name__contains=search_data)  # 搜索过滤

    # 2. 分页配置（每页5条）
    paginator = Paginator(query, 5)
    page_num = request.GET.get('page', 1)  # 拿页码，默认1

    # 3. 关键：容错处理（不管页码错成啥样，都返回有效页）
    try:
        c_page = paginator.page(page_num) 
          # 尝试拿指定页
    except PageNotAnInteger:  # 页码不是数字（比如?page=abc）
        c_page = paginator.page(1)  # 给第1页
    except EmptyPage:  # 页码超出范围（比如总1页，?page=2）
        c_page = paginator.page(paginator.num_pages)  # 给最后1页
    if paginator.num_pages <= 3:
            page_nums = paginator.page_range  # 所有页码
    else:
            page_nums = [1, 2, 3]

    # 4. 传数据到模板
    return render(request, 'room.html', {
        'c_page': c_page,
        'paginator': paginator,
        'search_data': search_data,
        "page_nums":page_nums
    })

def add_room(request):
    if request.method=="GET":
        titel="添加房间"
        form=roommodel()
        return render(request,"add_area.html",{"form":form,"titel":titel})
    form=roommodel(data=request.POST)
    if form.is_valid():
        data=request.session.get('info')
        if data==None:
            return HttpResponse("请先登录")
        name=data.get('name')
        logger.info('%s添加了一条房间记录%s'%(name,form.cleaned_data.get('room_name')))
        form.save()
        return redirect("/hostel/room/")
    return render(request,"add_area.html",{"form":form,"error":form.errors,'titel':"添加房间"})

def delete_room(request,id):
    page_num=request.GET.get("page","")
    try:
        page_num=int(page_num)
    except ValueError:
        page_num=1
    row_object=models.Room.objects.filter(id=id).first()
    if row_object==None:
        return HttpResponse("要删除的数据不存在")
    # data=request.session.get('info')
    # if data==None:
    #         return HttpResponse("请先登录")
    # name=data.get('name')
    # logger.info('%s删除了房间一条记录%s'%(name,row_object.room_name))
    row_object.delete()
    return redirect(f"/hostel/room/?page={page_num}")

class roomeditmodel(BootStrapModelForm):
    class Meta:
        model=models.Room
        fields=['people',"room_name","room_type"]
def edit_room(request,id):#编辑
    row_object=models.Room.objects.filter(id=id).first()
    if row_object==None:
        return HttpResponse("要修改的数据不存在")
    if request.method=="GET":
        title="修改-{}信息".format(row_object.room_name)
        form=roomeditmodel(instance=row_object)
        return render(request,"add_area.html",{"form":form,"titel":title})
    form=roomeditmodel(data=request.POST,instance=row_object)
    if form.is_valid():
        data=request.session.get('info')
        if data==None:
            return HttpResponse("请先登录")
        name=data.get('name')
        logger.info('%s更新了房间%s的记录'%(name,row_object.room_name))
        form.save()
        return redirect("/hostel/room/")
    return render(request,"add_area.html",{"form":form,"error":form.errors})

def add_all_room(request):#批量上传
    if request.method == 'GET':
        return render(request, 'add_all_area.html')

    excel_file = request.FILES.get('file')
    if not excel_file:
        return HttpResponse("请选择文件")

    wb = load_workbook(excel_file)
    sheet = wb.active
    errors = []  # 收集错误信息

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # 解析数据（按Excel列顺序：name, phone, depart_id, gender_id, permission_id, room_id, user_status_id）
            data = {
                'room_name': str(row[0]).strip() if row[0] else f"未知{idx}",
                "people":str(row[1]).strip() if row[1] else f"未知{idx}",
                'dorm':models.dorm.objects.get(name=row[2]),
                'gender': models.gender.objects.get(id=row[3]),
                'room_status':models.room_status.objects.get(id=row[4]),
                'room_perpe':str(row[5]).strip() if row[5] else f"未知{idx}" # 假设第8列是personne_id
            }
            models.Room.objects.create(**data)  # 解包创建
        except Exception as e:
            errors.append(f"第{idx}行错误：{str(e)}")  # 记录错误行

    # 处理结果
    if errors:
        return HttpResponse("导入失败：\n" + "\n".join(errors))
    
    logger.info(f"{request.session.get('info', {}).get('name')}批量添加人员")
    return redirect('/hostel/room/')

##批量删除
def delete_all_room(request):
    if request.method == 'POST':  # ① 判断请求是否为POST（只有提交表单时才会进入这个分支）
        ids = request.POST.getlist('ids')  # ② 获取前端勾选的人员ID列表（name="ids"的复选框值）
        if not ids:  # ③ 如果没有选中任何ID，给出错误提示
            messages.error(request, "请选择要删除的人员")
            return redirect('/hostel/room/')  # 重定向回人员列表页
        
        try:
            # ④ 批量删除选中的人员（根据ID列表过滤并删除）
            models.Room.objects.filter(id__in=ids).delete()
            messages.success(request, f"成功删除 {len(ids)} 条人员记录")  # ⑤ 成功提示
        except Exception as e:  # ⑥ 捕获删除过程中可能出现的异常（如外键关联错误）
            messages.error(request, f"删除失败：{str(e)}")
        
        return redirect('/hostel/room/')  # ⑦ 无论成功或失败，都重定向回人员列表页
    
    # 如果是GET请求（直接访问这个URL），也重定向回列表页
    return redirect('/hostel/room/')



def status(request):
    return render(request,"statusUI.html")


    
    
