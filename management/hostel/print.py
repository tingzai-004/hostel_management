from django.shortcuts import get_object_or_404, render,HttpResponse,redirect
from . import models
from hostel import models
from hostel.models import area ,dorm
from django import forms
from openpyxl import load_workbook
from hostel.bootstrap import BootStrapModelForm,BootStrapForm
import logging
from django.core.paginator import Paginator
import io
from .login import md5
import logging
from django.db.models import Count,Sum
import os
from django.conf import settings
from datetime import datetime
import random
import string

##导出
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
def get_random_str(length=8):
    characters = string.ascii_letters + string.digits
    return ''.join(random.choices(characters, k=length))

def write_to_excel1(data_list, file_path):
    if not data_list:
        workbook = openpyxl.Workbook()
        workbook.save(file_path)
        return
    
    # 中文表头映射
    header_mapping = {
        "id": "序号",
        "room": "房间号",
        "feetype": "费用类型",
        "amount": "缴纳金额",
        "status": "缴纳状态",
        "cdate": "创建日期",
        "check_in_people":"记录人"
    }
    headers = [header_mapping.get(key, key) for key in data_list[0].keys()]
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # 写入表头并设置样式
    for col_num, header_name in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header_name
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # 写入数据
    for row_num, row_data in enumerate(data_list, 2):
        for col_num, header_name in enumerate(headers, 1):
            original_key = list(data_list[0].keys())[col_num - 1]
            value = row_data.get(original_key, "")
            sheet.cell(row=row_num, column=col_num).value = value
    
    # 设置列宽
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].width = 20  # 调整为合适宽度
    
    workbook.save(file_path)

#费用记录打印
def peint(request):
    
    obj_fee=models.fee_record.objects.all().values()
    fee = list(obj_fee)
    
    excel_name = f"feerecord_{get_random_str()}.xlsx"
    path = os.path.join(settings.MEDIA_ROOT, excel_name)
    
    write_to_excel1(fee, path)
    return HttpResponse("导出成功")

##分摊记录打印
def write_to_excel2(data_list, file_path):
    if not data_list:
        workbook = openpyxl.Workbook()
        workbook.save(file_path)
        return
    
    # 中文表头映射
    header_mapping = {
        "id": "序号",
        "occupancyrecord_id":"名字",
        "feesharings": "分摊金额",
        "fee_type_id": "费用类型",
        "fee_record_id":"费用细则",
        "status": "缴纳状态",
        
        
    }
    headers = [header_mapping.get(key, key) for key in data_list[0].keys()]
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # 写入表头并设置样式
    for col_num, header_name in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header_name
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # 写入数据
    for row_num, row_data in enumerate(data_list, 2):
        for col_num, header_name in enumerate(headers, 1):
            original_key = list(data_list[0].keys())[col_num - 1]
            value = row_data.get(original_key, "")
            sheet.cell(row=row_num, column=col_num).value = value
    
    # 设置列宽
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].width = 20  # 调整为合适宽度
    
    workbook.save(file_path)

#分摊记录打印
def peint2(request):
    
    obj_fee=models.feesharing.objects.all().values()
    fee = list(obj_fee)
    
    excel_name = f"feesharing_{get_random_str()}.xlsx"
    path = os.path.join(settings.MEDIA_ROOT, excel_name)
    
    write_to_excel2(fee, path)
    return HttpResponse("导出成功")

    

