from django.shortcuts import render,HttpResponse,redirect
from . import models
from hostel import models
from hostel.models import area ,dorm
from django import forms
from openpyxl import load_workbook
from hostel.bootstrap import BootStrapModelForm
import logging
from django.core.paginator import Paginator,PageNotAnInteger,EmptyPage
from django.db.models import Sum,Count
import hashlib
from django.contrib import messages

def extract_number(text):
    # 定义中文数字和阿拉伯数字的对应关系
    num_map = {'一':1, '二':2, '三':3, '四':4, '五':5, 
               '六':6, '七':7, '八':8, '九':9, '十':10}
    
    # 遍历字符串中的每一个字
    for char in text:
        # 如果这个字是中文数字，直接返回对应的阿拉伯数字
        if char in num_map:
            return num_map[char]
    
    # 如果一个数字也没找到，返回 0 或其他你想要的默认值
    return 0

def md5(data_string):
    if data_string is None:
        # 可以选择返回一个默认值，或者抛出更明确的异常
        raise ValueError("Cannot encode None value")
    obj = hashlib.md5()
    obj.update(data_string.encode('utf-8'))
    return obj.hexdigest()

logger=logging.getLogger('management')
class personmodel(BootStrapModelForm):
    class Meta:
        model=models.Person
        fields=['name','depart','gender','phone','userstatus','personne','permission',"password"]
    # def clean(self):
    #     """自定义验证：检查房间是否已满"""
    #     cleaned_data = super().clean()##什么意思？
    #     room = cleaned_data.get('room')
        
    #     if not room:
    #         return cleaned_data  # 房间未选择时不验证
        
    #     # 提取额定人数（从 room_perpe 如"6R"中提取数字）
    #     try:
    #         rated_people = int(''.join(filter(str.isdigit, room.room_perpe)))
    #     except:
    #         raise forms.ValidationError(f"房间{room.room_name}属性格式错误（应为类似'6R'）")
        
    #     # 计算当前入住人数
    #     current_people =models.Person.objects.filter(room=room).count()
        
    #     # 检查是否超员（如果是新增，当前人数+1是否超过额定）
    #     if current_people >= rated_people:
    #         raise forms.ValidationError(f"房间{room.room_name}已住满（额定{rated_people}人）")
        
    #     return cleaned_data
        
def person(request):
    search_data = request.GET.get('q', '')
    query = models.Person.objects.all()
    if search_data:
        query = query.filter(name__contains=search_data)  # 搜索过滤

    # 2. 分页配置（每页5条）
    paginator = Paginator(query, 5)
    page_num = request.GET.get('page', 1)  # 拿页码，默认1

    # 3. 关键：容错处理（不管页码错成啥样，都返回有效页）
    try:
        c_page = paginator.page(page_num)
          # 尝试拿指定页
    except PageNotAnInteger:  # 页码不是数字（比如?page=abc）
        c_page = paginator.page(1)  # 给第1页
    except EmptyPage:  # 页码超出范围（比如总1页，?page=2）
        c_page = paginator.page(paginator.num_pages)  # 给最后1页
    if paginator.num_pages <= 3:
            page_nums = paginator.page_range  # 所有页码
    else:
            page_nums = [1, 2, 3] 

    # 4. 传数据到模板
    return render(request, 'person.html', {
        'c_page': c_page,
        'paginator': paginator,
        'search_data': search_data,
        "page_nums":page_nums
    })

class addpersonmodel(BootStrapModelForm):
    bootstrap_exclude=["money"]
    class Meta:
        model=models.Person
        fields=['name','depart','gender','phone','userstatus','personne','permission',"password"]
    def password_clean(self):
        pwd=self.cleaned_data.get("password")
        pwd=md5(pwd)  
        return pwd
    # def clean(self):
    #     """自定义验证：检查房间是否已满"""
    #     cleaned_data = super().clean()##什么意思？
    #     room = cleaned_data.get('room')
        
    #     if not room:
    #         return cleaned_data  # 房间未选择时不验证
        
    #     # 提取额定人数（从 room_perpe 如"6R"中提取数字）
    #     try:
    #         rated_people = room.room_type.bed_count
    #     except:
    #         raise forms.ValidationError(f"房间{room.room_name}属性格式错误（应为类似'6R'）")
        
    #     # 计算当前入住人数
    #     current_people =models.Person.objects.filter(room=room).count()
        
    #     # 检查是否超员（如果是新增，当前人数+1是否超过额定）
    #     if current_people >= rated_people:
    #         raise forms.ValidationError(f"房间{room.room_name}已住满（额定{rated_people}人）")
        
    #     return cleaned_data      

    
   
def add_person(request):
    if request.method=="GET":
        form=addpersonmodel()
        
        return render(request,"add_area.html",{"form":form,"titel":"添加人员信息"})
    form=addpersonmodel(data=request.POST)
    if form.is_valid():
        # 保存前先检查房间是否已满（Form的clean方法已验证，但需同步更新房间状态）
        form.save()
        return redirect("/hostel/person/")
    return render(request,"add_area.html",{'form':form,"error":form.errors,"titel":"添加人员信息"})
    #     new_person = form.save()
    #     room = new_person.room
        
    #     # 计算当前房间人数
    #     rated_people =room.room_type.bed_count
    #     current_people =models. Person.objects.filter(room=room).count()
        
    #     # 更新房间人数和状态
    #     room.people = current_people
    #     if current_people >= rated_people:
    #         full_status = models.room_status.objects.get(name="住满")
    #         room.room_status = full_status
    #         messages.success(request, f"添加成功！房间{room.room_name}已住满")
    #     else:
    #         messages.success(request, f"添加成功！房间当前{current_people}/{rated_people}人")
    #     logger.info(f"{request.session.get('info',{}).get('name')}添加了一条人员记录")
    #     room.save()
    #     return redirect("/hostel/person/")  # 跳转到人员列表页（避免重复提交）
    # else:
    #     # 表单验证失败（含房间已满的错误），将错误传递到模板
    #     messages.error(request, "表单验证失败或房间人数已满")
    #     return render(request, "add_area.html", {"form": form, "error": form.errors, "titel": "添加人员信息"})

##删除更新
def delete_person(request,id):
    page_num = request.GET.get("page", "1")
    try:
        page_num = int(page_num)
    except ValueError:
        page_num = 1  # 默认为第1页
    
    # 校验人员是否存在
    person = models.Person.objects.filter(id=id).first()
    if not person:
        return HttpResponse("找不到数据")
    # room=person.room
    # room_name=room.room_name
    try:
        person.delete()
        logger.info(f"{request.session.get('info',{}).get('name')}删除了一条人员记录")
        messages.success(request,"删除成功")
        return redirect(f"/hostel/person?page={page_num}")
    except Exception as e:
        messages.error(request,"删除失败，错误信息：{}".format(e))

    # room_person=models.Person.objects.filter(room=room).aggregate(total=Count("id"))["total"] or 0
    # rated_people = room.room_type.bed_count
    # # records=models.fee_record.objects.get(room=room)
    # # sharing=round(records.amount/room_person,2) if room_person>0 else 0
    
    # room.people=room_person

    # if room_person>=rated_people:
    #     full_status=models.room_status.objects.get(name="住满")
    #     room.room_status=full_status
    #     messages.success(request,f"删除成功，但{room_name}仍住满")
    # else:
    #     not_status=models.room_status.objects.get(name="可住")
    #     room.room_status=not_status
    #     messages.success(request,f"删除成功，{room_name}剩余{room_person}/{rated_people}")
    # logger.info(f"{request.session.get('info',{}).get('name')}删除了一条人员记录")
    # room.save()

    
    return redirect(f"/hostel/person/?page={page_num}")






def edit_person(request,id):#编辑
    row_object=models.Person.objects.filter(id=id).first()
    if row_object==None:
        return HttpResponse("要修改的数据不存在")
    if request.method=="GET":
        title="修改-{}".format(row_object.name)
        form=personmodel(instance=row_object)
        return render(request,"add_area.html",{"form":form,"titel":title})
    form=personmodel(data=request.POST,instance=row_object)
    if form.is_valid():
        data=request.session.get('info')
        if data==None:
            return HttpResponse("请先登录")
        name=data.get('name')
        logger.info('%s更新了人员表%s的信息'%(name,row_object.name))
        form.save()
        return redirect("/hostel/person/")
    return render(request,"add_area.html",{"form":form,"error":form.error,"titel":title})

def add_all_person(request):
    if request.method == 'GET':
        return render(request, 'add_all_area.html')

    excel_file = request.FILES.get('file')
    if not excel_file:
        return HttpResponse("请选择文件")

    wb = load_workbook(excel_file)
    sheet = wb.active
    errors = []  # 收集错误信息

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # 解析数据（按Excel列顺序：name, phone, depart_id, gender_id, permission_id, room_id, user_status_id）
            data = {
                'name': str(row[0]).strip() if row[0] else f"未知{idx}",
                'phone': str(row[1]).strip() if row[1] else "",
                # 外键实例查询（get_or_create：不存在则报错，确保数据有效）
                'depart': models.Depart.objects.get(id=row[2]),
                'gender': models.gender.objects.get(id=row[3]),
                'permission':models. Permission.objects.get(id=row[4]),
                'room': models.Room.objects.get(id=row[5]),
                'userstatus':models. userstatus.objects.get(id=row[6]),
                'personne':models. Personne.objects.get(id=row[7])  # 假设第8列是personne_id
            }
            models.Person.objects.create(**data)  # 解包创建
        except Exception as e:
            errors.append(f"第{idx}行错误：{str(e)}")  # 记录错误行

    # 处理结果
    if errors:
        return HttpResponse("导入失败：\n" + "\n".join(errors))
    
    logger.info(f"{request.session.get('info', {}).get('name')}批量添加人员")
    return redirect('/hostel/person/')
from django.contrib import messages

##批量删除
def delete_all_person(request):
    # 仅处理POST请求
    if request.method != 'POST':
        return redirect('person_list') # 或 HttpResponse("方法不允许")

    # 1. 验证用户登录状态
    name = request.session.get("info")
    if not name:
        return HttpResponse("请登录")
    admin_name = name.get("name")

    # 2. 获取并验证要删除的人员ID列表
    ids = request.POST.getlist('ids')
    if not ids:
        messages.error(request, "请选择要删除的人员")
        return redirect('/hostel/person/')

    try:
        persons_to_delete = models.Person.objects.filter(id__in=ids).select_related('room')
        
        # 收集所有涉及到的房间ID，用set去重
        affected_room_ids = {person.room.id for person in persons_to_delete}

        # 步骤2: 执行删除操作
        deleted_count, _ = models.Person.objects.filter(id__in=ids).delete()
        logger.info(f"管理员 {admin_name} 删除了 {deleted_count} 条人员记录。")

        # 步骤3: 遍历所有受影响的房间，更新房间人数和状态
        for room_id in affected_room_ids:
            room = models.Room.objects.get(id=room_id)
            
            # 重新计算当前房间的人数
            current_person_count = models.Person.objects.filter(room=room).count()
            
            # 更新房间人数
            room.people = current_person_count
            room.save()

            # 更新房间状态
            # 假设 room.room_perpe 是字符串类型的额定人数，例如 "4"
            try:
                rated_people = int(''.join(filter(str.isdigit, room.room_perpe)))
            except (ValueError, TypeError):
                rated_people = 0 # 或根据业务逻辑处理异常情况
            
            if current_person_count >= rated_people and rated_people > 0:
                full_status = models.room_status.objects.get(name="住满")
                room.room_status = full_status
                messages.success(request, f"删除成功，房间 {room.room_name} 状态更新为 '住满'。")
            else:
                not_full_status = models.room_status.objects.get(name="可住")
                room.room_status = not_full_status
                messages.success(request, f"删除成功，房间 {room.room_name} 状态更新为 '可住'。当前人数: {current_person_count}/{rated_people}。")
            
            room.save() # 再次保存状态变更

           
            unpaid_fee_records = models.fee_record.objects.filter(room=room)

            for fee_rec in unpaid_fee_records:
                # 获取当前房间内需要分摊此费用的所有人员
                # 这里假设只要是当前房间的人都需要分摊
                current_residents = models.Person.objects.filter(room=room)
                
                # 获取这些人员的住宿记录(occupancyrecord)
                # 注意：这里简化了逻辑，假设一个人在一个房间只有一条有效的住宿记录
                resident_ids = current_residents.values_list('id', flat=True)
                occupancy_records = models.occupancyrecord.objects.filter(user_id__in=resident_ids, room=room)

                num_residents = occupancy_records.count()

                if num_residents > 0:
                    # 计算单人分摊金额 (总金额 / 人数)
                    # 使用 Decimal 进行精确计算
                    from decimal import Decimal, ROUND_HALF_UP
                    total_fee = fee_rec.amount
                    individual_share = (total_fee / Decimal(num_residents)).quantize(Decimal('0.00'), rounding=ROUND_HALF_UP)
                    
                    # 更新或创建每个居民的 feesharing 记录
                    for occ_rec in occupancy_records:
                        fs, created = models.feesharing.objects.update_or_create(
                            fee_record=fee_rec,
                            occupancyrecord=occ_rec,
                            defaults={
                                'feesharings': str(individual_share), # 假设字段是CharField，需要转换
                                'fee_type': fee_rec.feetype, # 关联费用类型
                                'start_date': fee_rec.cdate, # 假设开始日期为费用创建日期
                                'end_date': fee_rec.cdate, # 这里可能需要更复杂的日期逻辑
                                'status': '0' # 标记为未缴纳
                            }
                        )
            # --- 费用重新计算结束 ---

        messages.success(request, f"成功删除 {deleted_count} 条人员记录，并已更新相关房间状态和费用分摊。")

    except models.Person.DoesNotExist:
        messages.error(request, "所选人员不存在或已被删除。")
    except models.room_status.DoesNotExist:
        messages.error(request, "房间状态配置错误，请联系管理员。")
    except Exception as e:
        # 捕获所有其他异常，方便调试
        messages.error(request, f"删除操作失败：{str(e)}")
        logger.error(f"删除人员时发生错误: {e}", exc_info=True)

    return redirect('/hostel/person/')
#住宿记录
class occupancyrecordmodel(BootStrapModelForm):
    bootstrap_exclude=["discount"]
    class Meta:
        model=models.occupancyrecord
        fields=['user','room','check_in_date','status','discount',"date_limit"] 
    def clean(self):
        """自定义验证：检查房间是否已满"""
        cleaned_data = super().clean()##什么意思？
        room = cleaned_data.get('room')
        
        if not room:
            return cleaned_data  # 房间未选择时不验证
        
        # 提取额定人数（从 room_perpe 如"6R"中提取数字）
        try:
            rated_people =room.room_type.bed_count
        except:
            raise forms.ValidationError(f"房间{room.room_name}属性格式错误（应为类似'6R'）")
        
        # 计算当前入住人数
        current_people =models.occupancyrecord.objects.filter(room=room,status=0).count()
        
        # 检查是否超员（如果是新增，当前人数+1是否超过额定）
        if current_people >= rated_people:
            raise forms.ValidationError(f"房间{room.room_name}已住满（额定{rated_people}人）")
        
        return cleaned_data      

def occupancyrecord(request):
    search_data = request.GET.get('q', '')
    query = models.occupancyrecord.objects.filter(status="0")
    if search_data:
        query = query.filter(user__name__contains=search_data)  # 搜索过滤

    # 2. 分页配置（每页5条）
    paginator = Paginator(query, 5)
    page_num = request.GET.get('page', 1)  # 拿页码，默认1

    # 3. 关键：容错处理（不管页码错成啥样，都返回有效页）
    try:
        c_page = paginator.page(page_num)
        if paginator.num_pages <= 3:
            page_nums = paginator.page_range  # 所有页码
        else:
            page_nums = [1, 2, 3]  # 只显示前3页  # 尝试拿指定页
    except PageNotAnInteger:  # 页码不是数字（比如?page=abc）
        c_page = paginator.page(1)  # 给第1页
    except EmptyPage:  # 页码超出范围（比如总1页，?page=2）
        c_page = paginator.page(paginator.num_pages)  # 给最后1页
    if paginator.num_pages <= 3:
            page_nums = paginator.page_range  # 所有页码
    else:
            page_nums = [1, 2, 3] 

    # 4. 传数据到模板
    return render(request, 'occupancyrecord.html', {
        'c_page': c_page,
        'paginator': paginator,
        'search_data': search_data,
        "page_nums":page_nums,
    })
   

def add_occupancyrecord(request):
    if request.method == "GET":
        form = occupancyrecordmodel()
        return render(request, "add_area.html", {"form": form, 'title': "添加住宿记录"})

    # 如果是 POST 请求
    form = occupancyrecordmodel(data=request.POST)
    if form.is_valid():
        person_instance = form.cleaned_data.get('user') 
        user_id = person_instance.id
        if models.occupancyrecord.objects.filter(user_id=user_id, status="0").exists():
            form.add_error("user", "该员工已有未退房的入住记录，不能重复添加。")
            return render(request, "add_area.html", {"form": form, "title": "添加住宿记录"})
        
        # 保存记录
        new_record = form.save()
        room = new_record.room
        current_people = models.occupancyrecord.objects.filter(room=room, status="0").count()
        person=models.Person.objects.get(id=user_id)
        person.room=room
        person.save()
        # 更新房间信息
        room.people = current_people
        relate_people=room.room_type.bed_count
        print("relate_people",relate_people)
        # 假设 room_perpe 是一个整数字段
        
        if current_people >= relate_people:
            full_status = models.room_status.objects.get(name="住满")
            room.room_status = full_status
            messages.success(request, f"添加成功！房间 {room.room_name} 已住满。")
        else:
            messages.success(request, f"添加成功！房间 {room.room_name} 当前 {current_people}/{room.room_type.bed_count} 人。") 
        logger.info(f"{request.session.get('info', {}).get('name')} 添加了一条入住记录 (员工: {user_id} - {person_instance.name}, 房间: {room.room_name})")
        room.save()
        return redirect("/hostel/occupancyrecord/")
    else:
        # 表单验证失败
        messages.error(request, "表单验证失败，请检查输入信息。")
        return render(request, "add_area.html", {"form": form, "title": "添加住宿记录"})
    

def delete_occupancyrecord(request,id):
    page_num=request.GET.get("page","")
    page_num=int(page_num)
    person=models.occupancyrecord.objects.filter(id=id).first()
    
    if not person:
        return HttpResponse("找不到数据")
    room=person.room
    room_name=room.room_name

    person.delete()

    room_person=models.occupancyrecord.objects.filter(room=room).aggregate(total=Count("id"))["total"] or 0
    rated_people = room.room_type.bed_count

    room.people=room_person

    if room_person>=rated_people:
        full_status=models.room_status.objects.get(name="住满")
        room.room_status=full_status
        messages.success(request,f"删除成功，但{room_name}仍住满")
    else:
        not_status=models.room_status.objects.get(name="可住")
        room.room_status=not_status
        messages.success(request,f"删除成功，{room_name}剩余{room_person}/{rated_people}")
    logger.info(f"{request.session.get('info',{}).get('name')}删除了一条住宿记录")
    room.save()   
    return redirect(f"/hostel/occupancyrecord/?page={page_num}")

class editoccupancyrecordmodel(BootStrapModelForm):
    class Meta:
        model=models.occupancyrecord
        fields=['user','room'] 


def edit_occupancyrecord(request,id):#编辑
    row_object=models.occupancyrecord.objects.filter(id=id).first()
    if row_object==None:
        return HttpResponse("要修改的数据不存在")
    if request.method=="GET":
        title="修改-{}".format(row_object.user.name)
        form=editoccupancyrecordmodel(instance=row_object)
        return render(request,"add_area.html",{"form":form,"titel":title})
    form=editoccupancyrecordmodel(data=request.POST,instance=row_object)
    if form.is_valid():
        data=request.session.get('info')
        if data==None:
            return HttpResponse("请先登录")
        name=data.get('name')
        logger.info('%s了楼栋数更新了%s的住宿记录'%(name,row_object.user))
        form.save()
        return redirect("/hostel/occupancyrecord/")
    return render(request,"add_area.html",{"form":form,"error":form.errors,"titel":title})




from datetime import datetime
def parse_date(date_value):
    if not date_value:  # 空值直接返回None
        return None
    
    # 情况1：如果已经是datetime对象，直接转为date
    if isinstance(date_value, datetime):
        return date_value.date()  # 提取date部分（去掉时间）
    
    # 情况2：如果是字符串，尝试解析
    if isinstance(date_value, str):
        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"]:
            try:
                return datetime.strptime(date_value, fmt).date()
            except ValueError:
                continue
    
    # 其他情况（非字符串也非datetime对象）
    raise ValueError(f"无法解析日期：{date_value}（类型：{type(date_value)})")



def add_all_occupancyrecord(request):
    if request.method == 'GET':
        return render(request, 'add_all_area.html')

    excel_file = request.FILES.get('file')
    if not excel_file:
        return HttpResponse("请选择文件")

    wb = load_workbook(excel_file)
    sheet = wb.active
    errors = []  # 收集错误信

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        try:
            # 1. 解析用户（外键：通过姓名关联Person）
            user_name = row[0]
            if not user_name:
                errors.append(f"第{idx}行错误：用户名为空，无法关联")
                continue  # 跳行
            user = models.Person.objects.filter(name=user_name).first()
            if not user:
                errors.append(f"第{idx}行错误：用户 '{user_name}' 不存在")
                continue  # 跳行

            # 2. 解析房间（外键：通过房间名称关联Room）
            room_name = row[1]
            if not room_name:
                errors.append(f"第{idx}行错误：房间名称为空，无法关联")
                continue  # 跳行
            room = models.Room.objects.filter(id=room_name).first()
            if not room:
                errors.append(f"第{idx}行错误：房间 '{room_name}' 不存在")
                continue  # 跳行

            # 3. 解析入住日期（调用辅助函数，处理空值和格式）
            try:
                check_in_date = parse_date(row[2])
            except ValueError as e:
                errors.append(f"第{idx}行错误：入住日期解析失败 - {str(e)}")
                continue  # 跳行

            # 4. 解析退房日期（同理）
            try:
                check_out_date = parse_date(row[3])
            except ValueError as e:
                errors.append(f"第{idx}行错误：退房日期解析失败 - {str(e)}")
                continue  # 跳行

            # 5. 解析状态（非空校验）
            status = row[4]
            if not status:
                errors.append(f"第{idx}行错误：状态为空")
                continue  # 跳行
            status = str(status).strip()  # 转为字符串并去空格

            # 6. 创建入住记录
            models.occupancyrecord.objects.create(
                user=user,
                room=room,
                check_in_date=check_in_date,
                check_out_date=check_out_date,
                status=status
            )

        except Exception as e:
            errors.append(f"第{idx}行错误：{str(e)}")  # 捕获其他意外错误

    # 处理结果
    if errors:
        return HttpResponse("导入失败：\n" + "\n".join(errors))
    
    logger.info(f"{request.session.get('info', {}).get('name')}批量添加入住记录")
    return redirect('/hostel/occupancyrecord/')

def delete_all_occupancyrecord(request):
    if request.method == 'POST':
        # 获取勾选的person ID列表（前端表单name为"ids"）
        ids = request.POST.getlist('ids')
        if not ids:
            messages.error(request, "请选择要删除的人员")
            return redirect('/hostel/occupancyrecord/')
        deleted_count, _ = models.occupancyrecord.objects.filter(id__in=ids).delete()
        logger.info(f"{request.session.get('info', {}).get('name')}批量删除 {deleted_count} 条入住记录")
        messages.success(request, f"成功删除 {deleted_count} 条入住记录")   
        return redirect('/hostel/occupancyrecord/')
    else:
        return HttpResponse("方法不允许")

        
        

class depmodel(BootStrapModelForm):
    class Meta:
        model=models.Depart
        fields=['depart_title','depart_number']

def dep(request):
    search_data = request.GET.get('q', '')
    query = models.Depart.objects.all()
    if search_data:
        query = query.filter(name__contains=search_data)  # 搜索过滤

    # 2. 分页配置（每页5条）
    paginator = Paginator(query, 5)
    page_num = request.GET.get('page', 1)  # 拿页码，默认1

    # 3. 关键：容错处理（不管页码错成啥样，都返回有效页）
    try:
        c_page = paginator.page(page_num)  # 尝试拿指定页
    except PageNotAnInteger:  # 页码不是数字（比如?page=abc）
        c_page = paginator.page(1)  # 给第1页
    except EmptyPage:  # 页码超出范围（比如总1页，?page=2）
        c_page = paginator.page(paginator.num_pages)  # 给最后1页

    # 4. 传数据到模板
    return render(request, 'dep.html', {
        'c_page': c_page,
        'paginator': paginator,
        'search_data': search_data
    })
def add_dep(request):
    if request.method=="GET":
        form=depmodel()
        
        return render(request,"add_area.html",{"form":form,'titel':"添加部门"})
    form=depmodel(data=request.POST)
    if form.is_valid():
        # data=request.session.get('info')
        # if data==None:
        #     return HttpResponse("请先登录")
        # name=data.get('name')
        # logger.info('%s添加了部门'%(name))
        form.save()
        return redirect("/hostel/dep/")
    return render(request,"add_area.html",{"form":form,"error":form.errors,'titel':"添加部门"})
def delete_dep(request,id):
    page_num=request.GET.get("page","")
    page_num=int(page_num)
    row_object=models.Depart.objects.filter(id=id).first()
    if row_object==None:
        return HttpResponse("数据不存在")
    data=request.session.get('info')
    if data==None:
        return HttpResponse("请先登录")
    name=data.get('name')
    logger.info('%s删除了部门数据'%(name))
    row_object.delete()
    return redirect(f"/hostel/dep/?page={page_num}")

def edit_dep(request,id):
    row_object=models.Depart.objects.filter(id=id).first()
    if row_object==None:
        return HttpResponse("要修改的数据不存在")
    if request.method=='GET':
        title="修改-{}".format(row_object.depart_title)
        form=depmodel(instance=row_object)
        return render(request,'add_area.html',{"form":form,"titel":title})
    form=depmodel(data=request.POST,instance=row_object)
    if form.is_valid():
        data=request.session.get('info')
        if data==None:
            return HttpResponse("请先登录")
        name=data.get('name')
        logger.info('%s更新了部门数据'%(name))
        form.save()
        return redirect("/hostel/dep/")
    return render(request,'add_area.html',{"form":form,"error":form.errors,"titel":title})

class user_typemodel(forms.ModelForm):
    class Meta:
        model=models.Personne
        fields=['name']

#用户类型
def user_type(request):
    search_data = request.GET.get('q', '')
    query = models.Personne.objects.all()
    if search_data:
        query = query.filter(name__contains=search_data)  # 搜索过滤

    # 2. 分页配置（每页5条）
    paginator = Paginator(query, 5)
    page_num = request.GET.get('page', 1)  # 拿页码，默认1

    # 3. 关键：容错处理（不管页码错成啥样，都返回有效页）
    try:
        c_page = paginator.page(page_num)  # 尝试拿指定页
    except PageNotAnInteger:  # 页码不是数字（比如?page=abc）
        c_page = paginator.page(1)  # 给第1页
    except EmptyPage:  # 页码超出范围（比如总1页，?page=2）
        c_page = paginator.page(paginator.num_pages)  # 给最后1页

    # 4. 传数据到模板
    return render(request, 'user_type.html', {
        'c_page': c_page,
        'paginator': paginator,
        'search_data': search_data,
       
    })
   

def add_user_type(request):
    if request.method == "GET":
        form = user_typemodel()
        
        return render(request, 'add_area.html', {"form": form,"titel":"添加用户类型"})
    
    # POST请求处理
    form = user_typemodel(data=request.POST)
    if form.is_valid():

        form.save()  # 通常验证通过后会保存数据
        return redirect('/hostel/user_type/')  # 重定向到成功页面
    else:
        # 验证失败，重新显示表单和错误信息
        return render(request, 'add_area.html', {"form": form})
def del_type(request,id):
    page_num=request.GET.get("page","")
    page_num=page_num
    row_object=models.Personne.objects.filter(id=id).first()
    if row_object==None:
       
        return HttpResponse("数据不存在")
    row_object.delete()
    return redirect(f'/hostel/user_type/?page={page_num}')

def update_type(request,id):
    row_object=models.Personne.objects.filter(id=id).first()
    if row_object==None:
        return HttpResponse("编辑的数据不存在")
    if request.method=="GET":
        
        form=user_typemodel(instance=row_object)
        return render(request,"add_area.html",{"form":form,"titel":"编辑用户类型"})
    form=user_typemodel(data=request.POST,instance=row_object)
    if form.is_valid():
        form.save()
        return redirect("/hostel/area/")
    return render(request,'add_area.html',{"form":form,"titel":"编辑用户类型"})

class permissionmodel(BootStrapModelForm):
    class Meta:
        model=models.Permission
        fields="__all__"


def permission(request):
    search_data = request.GET.get('q', '')
    query = models.Permission.objects.all()
    if search_data:
        query = query.filter(name__contains=search_data)  # 搜索过滤

    # 2. 分页配置（每页5条）
    paginator = Paginator(query, 5)
    page_num = request.GET.get('page', 1)  # 拿页码，默认1

    # 3. 关键：容错处理（不管页码错成啥样，都返回有效页）
    try:
        c_page = paginator.page(page_num)  # 尝试拿指定页
    except PageNotAnInteger:  # 页码不是数字（比如?page=abc）
        c_page = paginator.page(1)  # 给第1页
    except EmptyPage:  # 页码超出范围（比如总1页，?page=2）
        c_page = paginator.page(paginator.num_pages)  # 给最后1页

    # 4. 传数据到模板
    return render(request, 'permission.html', {
        'c_page': c_page,
        'paginator': paginator,
        'search_data': search_data
    })

def add_permission(request):
    if request.method == "GET":
        form = permissionmodel()
        return render(request, 'add_area.html', {"form": form,'titel':"添加权限列表"})
    
    # POST请求处理
    form = permissionmodel(data=request.POST)
    if form.is_valid():

        form.save()  # 通常验证通过后会保存数据
        return redirect('/hostel/permission/')  # 重定向到成功页面
    else:
        # 验证失败，重新显示表单和错误信息
        return render(request, 'add_area.html', {"form": form,'titel':"添加权限列表"})
    
def del_permission(request,id):
    page_num=request.GET.get("page","")
    page_num=int(page_num)
    row_obj=models.Permission.objects.filter(id=id).first()
    if not row_obj:
        return HttpResponse("数据不存在")
    info=request.session.get("info")
    if info is None:
        return HttpResponse("登录")
    name=info.get("name")
    logger.info(f"{name}删除了一条权限列表信息")
    row_obj.delete()
    return redirect(f"/hostel/permission/?page={page_num}")

def edit_permission(request,id):
    row_obj=models.Permission.objects.filter(id=id).first()
    if not row_obj:
        return HttpResponse("数据不存在")
    if request.method=="GET":
        form=permissionmodel(instance=row_obj)
        return render(request,"add_area.html",{"form":form,'titel':"编辑权限列表"})
    form=permissionmodel(data=request.POST,instance=row_obj)
    if form.is_valid():
        form.save()
        return redirect("/hostel/permission/")
    return render(request,"add_area.html",{"form":form,"titel":"编辑权限列表"})



