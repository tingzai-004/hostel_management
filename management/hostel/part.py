from django.shortcuts import get_object_or_404, render,HttpResponse,redirect
from . import models
from hostel import models
from hostel.models import area ,dorm
from django import forms
from openpyxl import load_workbook
from hostel.bootstrap import BootStrapModelForm
import logging
from django.core.paginator import Paginator
import io
from .login import md5
import logging
from django.db.models import Count,Sum
import os
from django.conf import settings
from datetime import datetime
import random
import string
from decimal import Decimal
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

logger=logging.getLogger("management")

class ResetPasswordForm(forms.ModelForm):
    # 确认新密码字段（仅用于验证，不关联模型）
    confirm_password = forms.CharField(
        label="确认新密码",
        widget=forms.PasswordInput(render_value=True)  # 密码错误时保留输入
    )

    class Meta:
        model = models.Admin
        fields = ["password"]  # 仅需新密码字段（模型中存储密码的字段）
        labels = {
            "password": "新密码"  # 显示为“新密码”而非默认的“password”
        }
        widgets = {
            "password": forms.PasswordInput(render_value=True)  # 密码框，错误时保留输入
        }

    def clean_password(self):
        """验证新密码不能与原密码相同"""
        new_password = self.cleaned_data.get("password")  # 获取明文新密码
        new_password_md5 = md5(new_password)  # 加密新密码
        
        # 对比新密码加密后是否与数据库中原密码一致
        if self.instance.password == new_password_md5:
            raise forms.ValidationError("新密码不能与原密码相同")
        
        return new_password  # 返回明文，后续在save中统一加密

    def clean(self):
        """验证两次输入的新密码是否一致"""
        cleaned_data = super().clean()
        new_password = cleaned_data.get("password")
        confirm_password = cleaned_data.get("confirm_password")
        
        if new_password != confirm_password:
            raise forms.ValidationError("两次输入的密码不一致")
        
        return cleaned_data

    def save(self, commit=True):
        """保存时加密新密码"""
        admin = self.instance  # 当前登录用户实例
        admin.password = md5(self.cleaned_data["password"])  # 加密后更新密码
        if commit:
            admin.save()
        return admin

def admin_reset(request):
    # 验证登录状态
    info = request.session.get("info")
    if not info:
        return HttpResponse("请先登录")
    
    user_id = info.get("id")
    user_obj = models.Admin.objects.filter(id=user_id).first()
    if not user_obj:
        return HttpResponse("用户不存在")

    if request.method == "GET":
        # 初始化表单，关联当前用户实例（用于验证原密码）
        form = ResetPasswordForm()
        return render(request, "add_area.html", {
            "title": f"重置密码 - {user_obj.name}",
            "form": form
        })

    # 处理密码重置提交
    form = ResetPasswordForm(data=request.POST, instance=user_obj)
    if form.is_valid():
        form.save()  # 保存加密后的新密码
        logger.info(f"{user_obj.name} 重置了密码")
        request.session.flush()  # 清除session，强制重新登录
        return redirect("/login/")
    
    # 表单验证失败，回显错误
    return render(request, "add_area.html", {
        "title": f"重置密码 - {user_obj.name}",
        "form": form
    })

from decimal import Decimal, ROUND_HALF_UP
from django.shortcuts import render
from datetime import datetime
from decimal import Decimal
##报表
# ====================== 终极正确版：宿舍费用统计报表 ======================

        
    
class Usermodel(BootStrapModelForm):
    bootstrap_exclude=["img"]
    img=forms.CharField(max_length=20,label="头像")
    class Meta:
        model=models.user_img
        fields=["name","img"]
##修改背景
def show_image(request):
    # 查询最新上传的图片（按id倒序，取第一条）
    img = models.background.objects.order_by('-id').first()
    return render(request, 'background.html', {'img': img})

##导出
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def write_to_excel(data_list, file_path):
    if not data_list:
        workbook = openpyxl.Workbook()
        workbook.save(file_path)
        return
    
    # 中文表头映射
    header_mapping = {
        "id": "编号",
        "oc"
        "room": "房间号",
        "total": "总金额",
        "paid": "缴纳金额",
        "status": "缴纳状态",
        "date": "日期",
    }
    headers = [header_mapping.get(key, key) for key in data_list[0].keys()]
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # 写入表头并设置样式
    for col_num, header_name in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header_name
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # 写入数据
    for row_num, row_data in enumerate(data_list, 2):
        for col_num, header_name in enumerate(headers, 1):
            original_key = list(data_list[0].keys())[col_num - 1]
            value = row_data.get(original_key, "")
            sheet.cell(row=row_num, column=col_num).value = value
    
    # 设置列宽
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].width = 20  # 调整为合适宽度
    
    workbook.save(file_path)

def get_random_str(length=8):
    characters = string.ascii_letters + string.digits
    return ''.join(random.choices(characters, k=length))

def peint(request):
    obj_fee = models.fee.objects.all().values()
    fee = list(obj_fee)
    
    excel_name = f"students_{get_random_str()}.xlsx"
    path = os.path.join(settings.MEDIA_ROOT, excel_name)
    
    write_to_excel(fee, path)
    return HttpResponse("导出成功")
    
def table2(request):
    return render(request,"table2.html")


def table(request):
    from datetime import datetime
    from decimal import Decimal
    from django.db.models import Sum

    # 1. 获取用户选择的月份（没选就用当前月）
    month_str = request.GET.get('month')
    if month_str and len(month_str) == 7:  # 2025-10
        year, month = map(int, month_str.split('-'))
        print(year,month)
    else:
        today = datetime.today()
        year, month = today.year, today.month

    current_month = datetime(year, month, 1)

    # 2. 【重点】计算当月每个启用费用类型的总金额（水电房空调全对！）
    active_fees = []
    fee_types = models.feetype.objects.filter(status="启用").order_by('id')
    sharing_one= models.sharing.objects.filter(fee_record__date__year=year, date__month=month)
    room_fee=sharing_one.aggregate(total=Sum("room_fee"))["total"] or Decimal('0.00')
    room_fee=Decimal(room_fee)
    yuan_money=sharing_one.aggregate(total=Sum("yuan_money"))["total"] or Decimal('0.00')
    yuan_money=Decimal(yuan_money)
    ketizu_money=sharing_one.aggregate(total=Sum("ketizu_money"))["total"] or Decimal('0.00')
    ketizu_money=Decimal(ketizu_money)


    # 当月所有费用记录
    monthly_records = models.fee_record_one.objects.filter(
        date__year=year,
        date__month=month
    ).select_related('room__dorm__area')

    for ft in fee_types:
        total = Decimal('0.00')
        room_count = 0

        
        for rec in monthly_records:
                if not rec.room or not rec.room.dorm:
                    continue
                # 如果费用类型绑定了区域，只算该区域
                if ft.area and rec.room.dorm.area != ft.area:
                    continue
                amount = Decimal(rec.dynamic_fees.get(ft.name, '0'))
                if amount > 0:
                    total += amount
                    room_count += 1

        active_fees.append({
            'name': ft.name,
            'amount': total.quantize(Decimal('0.00')),
            'room_count': room_count,
        })

    # 3. 其他统计数据（总费用、已收、未收、住宿人数等）
    all_records = models.fee_record_one.objects.filter(date__year=year, date__month=month)
    
    total1 = all_records.aggregate(t=Sum('amount'))['t'] or Decimal('0')
    paid = all_records.filter(status='1').aggregate(t=Sum('amount'))['t'] or Decimal('0')
    unpaid = all_records.filter(status='0').aggregate(t=Sum('amount'))['t'] or Decimal('0')

    total1 = total1.quantize(Decimal('0.00'))
    paid = paid.quantize(Decimal('0.00'))
    unpaid = unpaid.quantize(Decimal('0.00'))

    rate = round(float(paid) / float(total1) * 100, 2) if total1 > 0 else 0
    rate_2 = round(float(unpaid) / float(total1) * 100, 2) if total1 > 0 else 0

    people = models.Room.objects.aggregate(p=Sum('people'))['p'] or 0
    room_sum = models.Room.objects.count()

    # 4. 费用分摊明细表格（原来的 sharing 数据）
    queryset = models.sharing.objects.select_related(
    'user', 'fee_record', 'fee_record__room', 'fee_record__room__dorm'
).order_by('-id')

    # 状态筛选（房间状态？缴费状态？这里按您模板写的是房间状态）
    room_id = request.GET.get('current_rooms')  # 您模板里 name="status"，但其实是房间ID
    if room_id and room_id.isdigit():
        queryset = queryset.filter(fee_record__room_id=room_id)

    paginator = Paginator(queryset, 5)
    page_number = request.GET.get('page')
    form = paginator.get_page(page_number)
    current = form.number
    total_pages = paginator.num_pages

    # 核心算法：始终显示当前页前后各1页，共3个（边界处理）
    if total_pages <= 3:
        display_pages = list(range(1, total_pages + 1))
    else:
        if current <= 2:
            display_pages = [1, 2, 3]
        elif current >= total_pages - 1:
            display_pages = [total_pages - 2, total_pages - 1, total_pages]
        else:
            display_pages = [current - 1, current, current + 1]
    try:
        c_page = paginator.page(page_number)
    except PageNotAnInteger:
        c_page = paginator.page(1)
    except EmptyPage:
        c_page = paginator.page(paginator.num_pages)
    
  

    # 5. 上下文传给模板
    context = {
        'active_fees': active_fees,
        'current_month': current_month,
        'room_fee' : room_fee,
        'yuan_money': yuan_money,
        'ketizu_money': ketizu_money,
        'total1': total1,
        'total_count': paid,
        'total_not': unpaid,
        'rate': rate,
        'rate_2': rate_2,
        'people': int(people),
        'room_sum': room_sum,
        'c_page': c_page,
        'current': current,
        'display_pages': display_pages,
        'form': form,
        'paginator': paginator,
        'rooms':models.Room.objects.all().order_by('room_name')  }

    return render(request, "table.html", context)


    
    